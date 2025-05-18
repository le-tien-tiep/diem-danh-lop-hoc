import cv2
import os
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing import image
from tensorflow.keras.preprocessing.image import img_to_array
import keras
import tkinter as tk
from tkinter import messagebox
import subprocess
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import threading
import tempfile

# Load model
model = keras.models.load_model("C:/Users/roxanne/Downloads/bai tap cua tiep/bai tiep nha/nhan dien mat/pp1.h5", compile=False)
label_map = {0: 'tronghia', 1: 'tienfall', 2: 'khoa', 3: 'Tiep', 4: 'vuong', 5:'vnghia'}
attendance_log = {}
cap = None
is_running_camera = False
last_seen_faces = {}

def recognize_face(face_img):
    face_img = cv2.resize(face_img, (128, 128))
    face_img = img_to_array(face_img)
    face_img = np.expand_dims(face_img / 255.0, axis=0)
    preds = model.predict(face_img)
    label_idx = np.argmax(preds)
    confidence = preds[0][label_idx]
    return label_map[label_idx], confidence

def process_camera():
    global cap, is_running_camera, attendance_log, last_seen_faces
    cap = cv2.VideoCapture(0)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    is_running_camera = True
    last_seen_faces = {}

    while is_running_camera:
        ret, frame = cap.read()
        now = datetime.now()
        current_faces = set()

        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face_roi = frame[y:y+h, x:x+w]
            name, conf = recognize_face(face_roi)

            if conf > 0.85:
                face_img = frame[y:y+h, x:x+w]
                with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp_file:
                    temp_image_path = tmp_file.name
                    cv2.imwrite(temp_image_path, face_img)

                    if name not in attendance_log:
                        attendance_log[name] = {
                            'Time In': now,
                            'Time Out': None,
                            'First Seen': now,
                            'Last Seen': now,
                            'IsPresent': True,
                            'ChuaDuTiet': "Không",
                            'Image Path': temp_image_path
                        }
                        print(f"{name} - Vào lúc: {now.strftime('%Y-%m-%d %H:%M:%S')}")
                    else:
                        attendance_log[name]['Last Seen'] = now
                        attendance_log[name]['IsPresent'] = True

                last_seen_faces[name] = now
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                cv2.putText(frame, name, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)

        exited_faces = set(attendance_log.keys()) - current_faces
        for exited_name in exited_faces:
            if attendance_log[exited_name]['IsPresent']:
                time_in = attendance_log[exited_name]['First Seen']
                time_difference = now - time_in
                if time_difference < timedelta(hours=4):
                    attendance_log[exited_name]['ChuaDuTiet'] = "Có"
                else:
                    attendance_log[exited_name]['Time Out'] = now
                    print(f"{exited_name} - Ra lúc: {now.strftime('%Y-%m-%d %H:%M:%S')}")
                attendance_log[exited_name]['IsPresent'] = False

        cv2.imshow("Face Attendance", frame)
        if cv2.waitKey(1) == 27:
            break

    if cap and cap.isOpened():
        cap.release()
    cv2.destroyAllWindows()
    save_attendance_data()

def start_attendance():
    global is_running_camera
    if not is_running_camera:
        threading.Thread(target=process_camera).start()
    else:
        messagebox.showinfo("Thông báo", "Camera đang chạy.")

def exit_app():
    global is_running_camera, cap, root
    is_running_camera = False
    if cap and cap.isOpened():
        cap.release()
    cv2.destroyAllWindows()
    root.destroy()

def save_attendance_data():
    global attendance_log
    wb = Workbook()
    ws = wb.active
    ws.title = "Diemdanh"
    ws.append(['Tên', 'Thời gian vào', 'Thời gian ra', 'Chưa đủ tiết', 'Hình ảnh'])

    temp_files_to_delete = []

    for idx, (name, log) in enumerate(attendance_log.items(), start=2):
        ws.cell(row=idx, column=1).value = name
        ws.cell(row=idx, column=2).value = log['Time In'].strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=idx, column=3).value = log.get('Time Out', "")
        ws.cell(row=idx, column=4).value = log['ChuaDuTiet']
        if os.path.exists(log['Image Path']):
            img = ExcelImage(log['Image Path'])
            img.height = 100
            img.width = 100
            ws.add_image(img, f"E{idx}")
            temp_files_to_delete.append(log['Image Path'])
        ws.row_dimensions[idx].height = 100

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25
    ws.column_dimensions[get_column_letter(5)].width = 80

    current_date = datetime.now().strftime("%d%m%Y")
    # Thay đổi đường dẫn này thành nơi bạn muốn lưu file
    save_directory = "C:/Users/roxanne/Downloads/bai tap cua tiep/bai tiep nha/nhan dien mat/"
    excel_filename = os.path.join(save_directory, f"diemdanhngay{current_date}.xlsx")
    wb.save(excel_filename)
    print(f"Đã lưu file điểm danh với hình ảnh vào {excel_filename}")

    for temp_file in temp_files_to_delete:
        try:
            os.remove(temp_file)
            print(f"Đã xóa file tạm: {temp_file}")
        except Exception as e:
            print(f"Lỗi khi xóa file tạm {temp_file}: {e}")

# Tạo cửa sổ giao diện chính
root = tk.Tk()
root.title("Hệ thống điểm danh bằng khuôn mặt")
root.geometry("640x480")

start_btn = tk.Button(root, text="Điểm danh", command=start_attendance,
                    font=("Arial", 16), bg="green", fg="white", width=20, height=2)
start_btn.pack(pady=50)

exit_btn = tk.Button(root, text="Thoát", command=exit_app,
                   font=("Arial", 16), bg="red", fg="white", width=20, height=2)
exit_btn.pack()

root.protocol("WM_DELETE_WINDOW", exit_app)

root.mainloop()
